# -*- coding: utf-8 -*-
"""
Created on Thu Oct 13 21:04:53 2022

@author: kflyn
"""
import openpyxl as xl
wb=xl.load_workbook("C:/Users\kflyn/Documents/B+/D_PIN_LIST.xlsx")
fuclist=['Function1','Function2','Function3','Function4','Function5','Function6','Function7','Function8']
#funcDic={'FUNC_0':1,'FUNC_1':2,'FUNC_2':3,'FUNC_3':4,'FUNC_4':5,'FUNC_5':6,'FUNC_6':7,'FUNC_7':8}
funcDic1={1:'FUNC_0',2:'FUNC_1',3:'FUNC_2',4:'FUNC_3',5:'FUNC_4',6:'FUNC_5',7:'FUNC_6',8:'FUNC_7'}
fuclist_for=['Function2','Function3','Function4','Function5','Function6','Function7','Function8']
ws4out = wb['Sheet4']
f = open('workfile.txt', 'w')
f1 = open('iomux_func.txt', 'w')
print(wb.sheetnames)
print("\nsheet titles: \n")
MyList=[]
for sheet in wb:
     print(sheet.title)
ws = wb['Sheet2']
ws.sheet_properties
ws5=wb['Sheet5']
rawi=0
coli=0
for col in ws.columns:
    if col[0].value in fuclist:
        print("col",coli,":",col[0].value)
        rawi=0
        for row in range(1,len(col)+1):
            print("rawi",rawi,col[rawi].value,ws4out.cell(rawi+1, coli+1).value)
            if col[rawi].value != None:
                ws4out.cell(rawi+1, coli+1).value=col[rawi].value
            else :
                ws4out.cell(rawi+1, coli+1).value="NON"
            rawi+=1
        coli+=1
    if col[0].value=="CHIP_PAD_NAME":
        for row in range(0,len(col)):
            ws5.cell(row+1,0+1).value=col[row].value
    if col[0].value in fuclist_for:
        strtmp="\n/*"+col[0].value+"*/\n"
        f1.write(strtmp)
        for row in range(1,len(col)):
            if (col[row].value != None) and (col[row].value not in MyList):
                strtmp="IOMUX_FUNC_"+col[row].value+",\n"
                strtmp=strtmp.upper()
                f1.write(strtmp)
                MyList.append(strtmp)
                
    if col[0].value=="FuncDefault":
        ws5.cell(0+1,2+1).value="FuncDefaultName"
        ws5.cell(0+1,1+1).value=col[0].value
        for row in range(1,len(col)):
            ws5.cell(row+1,1+1).value=col[row].value
            ws5.cell(row+1,2+1).value=funcDic1[col[row].value]
            ws5.cell(row+1,3+1).value=ws4out.cell(row+1,col[row].value).value
            
            
        #print(cell.value)
    
    rows=0
for row in ws4out.rows:
    cols=0
    rows+=1
    
    print(f'/*{ws4out.cell(rows,0+1).value:s}*/');
    f.write("/*"+ws4out.cell(rows,0+1).value+"*/\n")
    str0="{"
    for cell in row:
        if rows>=2:
            str0+=("IOMUX_FUNC_"+cell.value)
        else:
            str0+=(cell.value)
        str0=str0.upper()
        str0+=","
    str0+="},\n"
    print(str0)
    f.write(str0)
        
row_index=0

#for row in ws.iter_rows(min_row=1, max_col=30, max_row=14):
#    print("row%d:",row_index)
#    for cell in row:
#        print(cell.value)
#    row_index += 1
f.close()
f1.close()
wb.save("C:/Users/kflyn/Documents/B+/D_PIN_LIST.xlsx")
wb.close
    
    